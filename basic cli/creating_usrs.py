from openpyxl import *
from openpyxl.styles import Alignment,Font
import hashlib
wb = load_workbook("sample.xlsx")

ws= wb.active
#ws.title = "My Sheet"
#ws["A1"]= ""
#ws["B1"]= ""
#print(ws["B1"].value) get the value at each cell specified  
#print(wb.sheetnames) get the names of all the sheets we are working with
#wb.create_sheet("My Sheet 2")
#adding data to excel
#data = [["chetan","qwerty"],["verma","12345"],["abcd","0786"],["xyz","65548"],["ddddd","23232"],["dscs","2121"]] 

# for i in range(len(data)):
#     ws.append(data[i])
# for i in range(1,len(data)+1):
#     for j in range(1,3):
#         c = 64+j
#         c=chr(c)
#         ws[f"{c}{i}"]=""

# ws.merge_cells("A1:B1")
# ws.merge_cells("C1:D1")

# ws["A1"]="Username"
# ws["A1"].alignment = Alignment(horizontal='center')
# ws["C1"]="Password"
# ws["C1"].alignment= Alignment(horizontal='center')
# ws["A1"].font = Font(bold=True)
# ws["C1"].font = Font(bold=True)

total_rows = ws.max_row
usr = input()
pws = input()
l= [usr,pws]

ws.merge_cells(f"A{total_rows+1}:B{total_rows+1}")
ws.merge_cells(f"C{total_rows+1}:F{total_rows+1}")
c=0
for i in range(2,total_rows+1):
    if(ws[f"A{i}"].value==usr):
        print("the user name already taken try something else.")
        c=c+1
        break
if(c<=0):    
    ws[f"A{total_rows+1}"]=usr
    ws[f"C{total_rows+1}"]=hashlib.md5(pws.encode()).hexdigest()

# ws[f"{chr(64+1)}{total_rows+1}"]=usr
# ws[f"{chr(64+3)}{total_rows+1}"]=pws



wb.save("sample.xlsx")