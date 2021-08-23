from openpyxl import *
import hashlib
wb =load_workbook("sample.xlsx")
ws = wb.active
row = ws.max_row
user = input("Enter user: ")
pwd  = input("enter pass: ")

c=0
for i in range(2,row+1):
    if(user == ws[f"A{i}"].value and hashlib.md5(pwd.encode()).hexdigest() == ws[f"C{i}"].value):
        print("success")
        c=c+1
        break
if(c<=0):
    print("id or pass not correct")






